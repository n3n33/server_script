import openpyxl
import logging
import sys

from openpyxl import Workbook, workbook

def create_workbook():
  # workbook 생성
  wb = Workbook()
  # 시트명 가져옴
  ws = wb.active
  # 시트명 변경
  ws.title = 'testsheet'
  logging.info('Workbook sheet renamed')
  # 엑셀 저장
  wb.save('C:/Users/User/Desktop/excel_create_macro.xlsx')
  logging.info('Excel stored')
  wb.close()

def create_cells():
    logging.info('Workbook Load')
    wb = openpyxl.load_workbook('C:/Users/User/Desktop/excel_create_macro.xlsx')
    ws = wb['testsheet']
    logging.info('cell data input')
    ws['A1'] = 'testing'
    wb.save('C:/Users/User/Desktop/excel_create_macro.xlsx')


def input_data():
    logging.info('Workbook Load')
    wb = openpyxl.load_workbook('C:/Users/User/Desktop/excel_create_macro.xlsx')
    ws = wb['testsheet']
    logging.info('cell data input')
    i = 1
    for x in range(1, 20):
        for y in range(1,20):
            ws.cell(x, y).value = i
            i+=1
    wb.save('C:/Users/User/Desktop/excel_create_macro.xlsx')


# create_workbook()
# create_cells()
input_data()