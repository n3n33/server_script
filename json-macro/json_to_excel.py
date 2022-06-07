import logging
import sys
import json
import openpyxl
import pandas as pd

from openpyxl import Workbook

def loaded_json():
    with open('D:\github_task\server_script\json-macro\ip_addr.json', 'r') as f :
        json_data = json.load(f)
        print(json.dumps(json_data, indent="\t"))

def create_sheet():
    wb = Workbook()
    # 시트명 가져옴
    ws = wb.active
    # 시트명 변경
    ws.title = 'json_read'
    logging.info('Workbook sheet renamed')
    # 엑셀 저장
    wb.save('D:\github_task\server_script\json-macro\sample.xlsx')
    logging.info('Excel stored')
    wb.close()

def input_data():
    logging.info('Workbook Load')
    wb = openpyxl.load_workbook('D:\github_task\server_script\json-macro\sample.xlsx')
    ws = wb['json_read']
    logging.info('cell data input')
    ws.append(['test1','test2'])
    ws.append
    wb.save('D:\github_task\server_script\json-macro\sample.xlsx')

#loaded_json()
#create_sheet()
#input_data()

def using_pandas():
    with open('D:\github_task\server_script\json-macro\ip_addr.json', 'r') as f :
        json_data = json.load(f)
        #print(json.dumps(json_data, indent="\t"))
    #df_json = json.load(json_data)
    # ValueError: If using all scalar values, you must pass an index
    df = pd.DataFrame(json_data, index=[0])
    df.to_excel('test.xlsx')

using_pandas()